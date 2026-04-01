#!/usr/bin/env python3
"""
Script de teste para enviar formulário de manutenção e verificar geração do Excel
"""

import requests
import json
import time
from datetime import datetime
import base64

# URL da API
BASE_URL = "http://localhost:5000/api"

# Dados de teste do formulário completo
dados_teste = {
    # Página 1 - DADOS (OBRIGATÓRIA)
    "nome": "João Silva",
    "data": "2026-04-01",
    "hora": "14:30",
    "garagem": "Viação Metrópole AE Carvalho",
    "prefixo": "MAR-001",
    "id": "BUS-42",
    "tecnologia": "Articulado",
    "latitude": -23.5505,
    "longitude": -46.6333,
    "localizacao_completa": "Av. Paulista, 1000, Bela Vista, São Paulo, SP",
    "foto_frente_onibus": "data:image/jpeg;base64,/9j/4AAQSkZJRgABAQEAYABgAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCAABIAEDASIAAhEBAxEB/8QAFQABAQAAAAAAAAAAAAAAAAAAAAX/xAAUEAEAAAAAAAAAAAAAAAAAAAAA/8VAFQEBAQAAAAAAAAAAAAAAAAAAAAX/xAAUEQEAAAAAAAAAAAAAAAAAAAAA/9oADAMBAAIRAxEAPwCdABmX/9k=",
    
    # Página UCP
    "ucp_problemas": [],
    "ucp_acoes": [],
    "ucp_imagem_antes": None,
    "ucp_imagem_depois": None,
    
    # Página TDM
    "tdm_problemas": [],
    "tdm_acoes": [],
    "tdm_imagem_antes": None,
    "tdm_imagem_depois": None,
    
    # Página Switch
    "switch_problemas": [],
    "switch_acoes": [],
    "switch_imagem_antes": None,
    "switch_imagem_depois": None,
    
    # Página Antena
    "antena_problemas": [],
    "antena_acoes": [],
    "antena_imagem_antes": None,
    "antena_imagem_depois": None,
    
    # Observações
    "observacao": "Manutenção preventiva completa. Sistema funcionando normalmente.",
    "email": "joao.silva@marvel.com",
    "imagens_adicionais": []
}

def test_enviar_manutencao():
    """Testa envio de formulário de manutenção"""
    print("=" * 60)
    print("TESTE 1: Enviar Formulário de Manutenção")
    print("=" * 60)
    
    try:
        print(f"\n📤 Enviando formulário para: {BASE_URL}/manutencao/enviar")
        print(f"   Técnico: {dados_teste['nome']}")
        print(f"   Prefixo: {dados_teste['prefixo']}")
        print(f"   ID: {dados_teste['id']}")
        
        response = requests.post(
            f"{BASE_URL}/manutencao/enviar",
            json=dados_teste,
            timeout=30
        )
        
        print(f"\n✅ Status Code: {response.status_code}")
        resultado = response.json()
        
        print(f"✅ Resposta recebida:")
        print(f"   Status: {resultado.get('status')}")
        print(f"   Mensagem: {resultado.get('message')}")
        print(f"   Excel: {resultado.get('excel_name')}")
        print(f"   PDF: {resultado.get('pdf_gerado', 'N/A')}")
        
        return response.status_code == 200
        
    except requests.exceptions.ConnectionError:
        print("❌ ERRO: Não conseguiu conectar ao servidor!")
        print("   Certifique-se que o backend está rodando em http://localhost:5000")
        return False
    except Exception as e:
        print(f"❌ ERRO: {e}")
        return False

def test_download_excel():
    """Testa download do arquivo Excel"""
    print("\n" + "=" * 60)
    print("TESTE 2: Download do Arquivo Excel")
    print("=" * 60)
    
    try:
        print(f"\n📥 Solicitando download: {BASE_URL}/download-excel")
        
        response = requests.get(
            f"{BASE_URL}/download-excel",
            timeout=10
        )
        
        print(f"\n✅ Status Code: {response.status_code}")
        
        if response.status_code == 200:
            tamanho = len(response.content)
            print(f"✅ Arquivo recebido com sucesso!")
            print(f"   Tamanho: {tamanho} bytes")
            print(f"   Content-Type: {response.headers.get('content-type')}")
            
            # Salvar arquivo para verificação manual
            arquivo_local = "c:\\Marvel\\Manutenções_Marvel_TEST.xlsx"
            with open(arquivo_local, 'wb') as f:
                f.write(response.content)
            print(f"   Salvo em: {arquivo_local}")
            
            return True
        else:
            print(f"❌ Erro: {response.status_code}")
            print(f"   {response.text}")
            return False
            
    except requests.exceptions.ConnectionError:
        print("❌ ERRO: Não conseguiu conectar ao servidor!")
        return False
    except Exception as e:
        print(f"❌ ERRO: {e}")
        return False

def test_verificar_arquivo_excel():
    """Verifica se o arquivo Excel foi criado no backend"""
    print("\n" + "=" * 60)
    print("TESTE 3: Verificação do Arquivo Excel no Servidor")
    print("=" * 60)
    
    try:
        import os
        caminho_excel = "c:\\Marvel\\backend\\uploads\\Manutenções_Marvel.xlsx"
        
        if os.path.exists(caminho_excel):
            tamanho = os.path.getsize(caminho_excel)
            print(f"✅ Arquivo encontrado!")
            print(f"   Localização: {caminho_excel}")
            print(f"   Tamanho: {tamanho} bytes")
            
            # Tentar ler informações do Excel
            try:
                from openpyxl import load_workbook
                wb = load_workbook(caminho_excel)
                ws = wb.active
                
                print(f"   Planilha ativa: {ws.title}")
                print(f"   Linhas: {ws.max_row}")
                print(f"   Colunas: {ws.max_column}")
                
                # Mostrar primeiras linhas
                print(f"\n   Dados na planilha:")
                for row in ws.iter_rows(min_row=1, max_row=min(3, ws.max_row), values_only=True):
                    print(f"   {row}")
                
                return True
            except ImportError:
                print("   ⚠️  openpyxl não está instalado para verificação detalhada")
                return True
                
        else:
            print(f"❌ Arquivo NÃO encontrado em: {caminho_excel}")
            print(f"   Verifique se a manutenção foi enviada com sucesso")
            return False
            
    except Exception as e:
        print(f"❌ ERRO: {e}")
        return False

if __name__ == "__main__":
    print("\n")
    print("╔" + "═" * 58 + "╗")
    print("║" + " " * 10 + "TESTE DO SISTEMA - MARVEL MAINTENANCE" + " " * 11 + "║")
    print("║" + " " * 15 + "Enviando Formulário + Excel" + " " * 16 + "║")
    print("╚" + "═" * 58 + "╝")
    print(f"\nHora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"URL Base: {BASE_URL}")
    
    # Executar testes
    teste1 = test_enviar_manutencao()
    time.sleep(2)  # Aguardar processamento
    
    teste3 = test_verificar_arquivo_excel()
    
    teste2 = test_download_excel()
    
    # Resumo
    print("\n" + "=" * 60)
    print("RESUMO DOS TESTES")
    print("=" * 60)
    print(f"✓ Envio de Formulário: {'PASSOU' if teste1 else 'FALHOU'}")
    print(f"✓ Verificação Excel: {'PASSOU' if teste3 else 'FALHOU'}")
    print(f"✓ Download Excel: {'PASSOU' if teste2 else 'FALHOU'}")
    
    if teste1 and teste3 and teste2:
        print("\n🎉 TODOS OS TESTES PASSARAM!")
    else:
        print("\n⚠️  ALGUNS TESTES FALHARAM")
    
    print("=" * 60)
