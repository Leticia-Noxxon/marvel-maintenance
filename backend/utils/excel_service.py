import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink
from datetime import datetime
import os

def adicionar_hyperlink_imagem(cell, url_imagem):
    """
    Adiciona um hyperlink clicável em uma célula do Excel
    Ao clicar, abre a imagem no Google Drive
    """
    if url_imagem and url_imagem.startswith('http'):
        cell.value = "🖼️ Ver Imagem"
        cell.hyperlink = Hyperlink(ref=cell.coordinate, target=url_imagem)
        cell.font = Font(color="0563C1", underline="single")
    else:
        cell.value = "-"
    return cell

def atualizar_excel_manutencoes(dados_manutencao, caminho_excel='uploads/Manutenções_Marvel.xlsx'):
    """
    Atualiza um arquivo Excel com os dados de todas as manutenções
    Cria o arquivo se não existir
    """
    try:
        os.makedirs(os.path.dirname(caminho_excel), exist_ok=True)
        
        # Verificar se arquivo existe
        if os.path.exists(caminho_excel):
            wb = openpyxl.load_workbook(caminho_excel)
            ws = wb.active
            linha_inicio = ws.max_row + 1
        else:
            # Criar novo workbook com headers
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Registros"
            linha_inicio = 2
            
            # Headers
            headers = [
                'Data/Hora', 'Técnico', 'Prefixo', 'ID', 'Garagem', 'Tecnologia',
                'Localização Completa', 'Latitude', 'Longitude',
                'UCP - Problemas', 'UCP - Ações', 'UCP Foto Antes', 'UCP Foto Depois',
                'TDM - Problemas', 'TDM - Ações', 'TDM Foto Antes', 'TDM Foto Depois',
                'Switch - Problemas', 'Switch - Ações', 'Switch Foto Antes', 'Switch Foto Depois',
                'Antena - Problemas', 'Antena - Ações', 'Antena Foto Antes', 'Antena Foto Depois',
                'Observações', 'Foto Frente Ônibus'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Preparar dados para inserção
        nova_linha = [
            datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
            dados_manutencao.get('nome', ''),
            dados_manutencao.get('prefixo', ''),
            dados_manutencao.get('id', ''),
            dados_manutencao.get('garagem', ''),
            dados_manutencao.get('tecnologia', ''),
            dados_manutencao.get('localizacao_completa', ''),
            dados_manutencao.get('latitude', ''),
            dados_manutencao.get('longitude', ''),
            ', '.join(dados_manutencao.get('ucp_problemas', [])),
            ', '.join(dados_manutencao.get('ucp_acoes', [])),
            dados_manutencao.get('ucp_foto_antes_url', None),  # URL do Google Drive
            dados_manutencao.get('ucp_foto_depois_url', None),
            ', '.join(dados_manutencao.get('tdm_problemas', [])),
            ', '.join(dados_manutencao.get('tdm_acoes', [])),
            dados_manutencao.get('tdm_foto_antes_url', None),
            dados_manutencao.get('tdm_foto_depois_url', None),
            ', '.join(dados_manutencao.get('switch_problemas', [])),
            ', '.join(dados_manutencao.get('switch_acoes', [])),
            dados_manutencao.get('switch_foto_antes_url', None),
            dados_manutencao.get('switch_foto_depois_url', None),
            ', '.join(dados_manutencao.get('antena_problemas', [])),
            ', '.join(dados_manutencao.get('antena_acoes', [])),
            dados_manutencao.get('antena_foto_antes_url', None),
            dados_manutencao.get('antena_foto_depois_url', None),
            dados_manutencao.get('observacao', ''),
            dados_manutencao.get('foto_frente_onibus_url', None)
        ]

        # Inserir dados e adicionar hyperlinks para URLs
        for col, valor in enumerate(nova_linha, 1):
            cell = ws.cell(row=linha_inicio, column=col)
            
            # Se é uma URL de imagem, adicionar hyperlink
            if valor and isinstance(valor, str) and valor.startswith('http'):
                adicionar_hyperlink_imagem(cell, valor)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.value = valor
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Adicionar borda
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        
        for col in range(10, len(nova_linha) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        # Salvar
        wb.save(caminho_excel)
        print(f"✅ Excel atualizado: {caminho_excel}")
        return True, caminho_excel

    except Exception as e:
        print(f"❌ Erro ao atualizar Excel: {str(e)}")
        return False, str(e)


def gerar_resumo_excel(pasta_uploads='uploads'):
    """
    Gera um Excel com resumo executivo de todas as manutenções
    """
    try:
        caminho_excel = os.path.join(pasta_uploads, 'Relatório_Executivo_Marvel.xlsx')
        
        # Se não houver dados, criar vazio
        if not os.path.exists(os.path.join(pasta_uploads, 'Manutenções_Marvel.xlsx')):
            return False, "Nenhum dado disponível ainda"

        wb = openpyxl.Workbook()
        
        # Aba 1: Resumo Diário
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"
        
        ws_resumo['A1'] = "Manutenção Marvel - Relatório Executivo"
        ws_resumo['A1'].font = Font(bold=True, size=14)
        ws_resumo['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        
        # Dados resumidos
        ws_resumo['A4'] = "Total de Manutenções:"
        ws_resumo['B4'] = "=COUNTA(Registros!A:A)-1"  # Conta linhas menos header
        
        wb.save(caminho_excel)
        return True, caminho_excel

    except Exception as e:
        print(f"❌ Erro ao gerar relatório: {str(e)}")
        return False, str(e)
