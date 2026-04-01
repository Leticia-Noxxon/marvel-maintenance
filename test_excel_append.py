#!/usr/bin/env python3
"""
Teste de segunda submissão para confirmar append no Excel
"""

import requests
import time
from openpyxl import load_workbook

# Segunda submissão com dados diferentes
dados_teste_2 = {
    'nome': 'Maria Santos',
    'data': '2026-04-01',
    'hora': '15:45',
    'garagem': 'Viação Metrópole Itaim',
    'prefixo': 'MAR-002',
    'id': 'BUS-43',
    'tecnologia': 'Básico',
    'latitude': -23.5512,
    'longitude': -46.6340,
    'localizacao_completa': 'Rua Oscar Freire, 500, São Paulo, SP',
    'foto_frente_onibus': 'data:image/jpeg;base64,/9j/4AAQSkZJRgABAQEAYABgAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCAABIAEDASIAAhEBAxEB/8QAFQABAQAAAAAAAAAAAAAAAAAAAAX/xAAUEAEAAAAAAAAAAAAAAAAAAAAA/8VAFQEBAQAAAAAAAAAAAAAAAAAAAAX/xAAUEQEAAAAAAAAAAAAAAAAAAAAA/9oADAMBAAIRAxEAPwCdABmX/9k=',
    'ucp_problemas': [],
    'ucp_acoes': [],
    'ucp_imagem_antes': None,
    'ucp_imagem_depois': None,
    'tdm_problemas': [],
    'tdm_acoes': [],
    'tdm_imagem_antes': None,
    'tdm_imagem_depois': None,
    'switch_problemas': [],
    'switch_acoes': [],
    'switch_imagem_antes': None,
    'switch_imagem_depois': None,
    'antena_problemas': [],
    'antena_acoes': [],
    'antena_imagem_antes': None,
    'antena_imagem_depois': None,
    'observacao': 'Checagem de rotina realizada.',
    'email': 'maria@marvel.com',
    'imagens_adicionais': []
}

print("=" * 60)
print("TESTE DE APPEND - Segunda Submissão")
print("=" * 60)

print("\n📤 Enviando segunda manutenção...")
r = requests.post('http://localhost:5000/api/manutencao/enviar', json=dados_teste_2, timeout=30)
print(f"Status: {r.status_code}")
resp = r.json()
print(f"Mensagem: {resp['message']}")

# Verificar Excel agora
time.sleep(1)
print("\n📊 Verificando Excel atualizado...")

wb = load_workbook('c:\\Marvel\\backend\\uploads\\Manutenções_Marvel.xlsx')
ws = wb.active

print(f"Planilha: {ws.title}")
print(f"Total de linhas agora: {ws.max_row}")
print(f"Colunas: {ws.max_column}")

print(f"\n📋 Registros na planilha:")
for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
    if idx == 1:
        print(f"\nLinha {idx} (HEADER):")
        print(f"  {row[0:5]}")
    else:
        print(f"\nLinha {idx}:")
        print(f"  Data/Hora: {row[0]}")
        print(f"  Técnico: {row[1]}")
        print(f"  Prefixo: {row[2]}")
        print(f"  ID: {row[3]}")
        print(f"  Garagem: {row[4]}")
        print(f"  Tecnologia: {row[5]}")

if ws.max_row == 3:
    print("\n✅ SUCESSO! Excel possui HEADER + 2 registros")
    print("   Append funcionando corretamente!")
else:
    print(f"\n⚠️  Esperado 3 linhas (HEADER + 2 registros), mas tem {ws.max_row}")

print("\n" + "=" * 60)
